import sys
import os
import json
import tkinter as tk
from tkinter import simpledialog, messagebox
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from .layer_selector import selecionar_layers
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from ui.talhoes_parser import extrair_talhoes_por_proximidade,extrair_legenda_layers
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt
from PySide6.QtWidgets import QApplication
from ui.excel_viewer import PDFViewer
import subprocess
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage
from ui.imagem_utils import redimensionar_imagem, gerar_imagem_centrada, inserir_imagem
from openpyxl.worksheet.page import PageMargins
from ui.layout_dialog import ExtendedLayoutInfoDialog as LayoutInfoDialog
from PySide6.QtWidgets import QApplication, QMessageBox
import matplotlib.pyplot as plt
from matplotlib.patches import Circle, Polygon, Rectangle
import time
from collections import defaultdict
import math
from shapely.geometry import Polygon

MAX_DESENHISTA = 60  # Limite máximo para o nome do DESENHISTA


def resource_path(relative_path):
    """Retorna o caminho absoluto, mesmo quando empacotado com PyInstaller"""
    try:
        # Caso empacotado
        base_path = sys._MEIPASS
    except AttributeError:
        # Em desenvolvimento
        base_path = os.path.abspath(".")

    # Caminho completo real
    return os.path.normpath(os.path.join(base_path, relative_path))

def set_cell_value(ws, cell_coord, value):
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            anchor = merged_range.start_cell.coordinate
            ws[anchor].value = value
            return
    ws[cell_coord].value = value

def converter_excel_para_pdf_com_libreoffice(excel_path):
    try:
        libreoffice_path = os.path.join(os.getcwd(), "LibreOfficePortable", "App", "libreoffice", "program", "soffice.exe")
        print(f"📌 Caminho do LibreOffice: {libreoffice_path}")
        print(f"📌 Excel de entrada: {excel_path}")

        if not os.path.isfile(libreoffice_path):
            raise FileNotFoundError(f"❌ LibreOffice não encontrado em: {libreoffice_path}")

        excel_path = os.path.abspath(excel_path)
        output_dir = os.path.dirname(excel_path)
        pdf_path = excel_path.replace(".xlsx", ".pdf")
        print(f"📌 Diretório de saída: {output_dir}")
        print(f"📁 PDF alvo: {pdf_path}")

        # 💣 Remove PDF antigo se existir
        if os.path.exists(pdf_path):
            print("🧹 Removendo PDF anterior...")
            os.remove(pdf_path)

        command = [
            libreoffice_path,
            "--headless",
            "--nologo",
            "--convert-to", "pdf",
            "--outdir", output_dir,
            excel_path
        ]

        print("📤 Comando que será executado:")
        print(" ".join(command))

        creationflags = subprocess.CREATE_NO_WINDOW if sys.platform.startswith("win") else 0

        subprocess.run(command, check=True, creationflags=creationflags, timeout=30)

        # 🕒 Aguarda a criação do novo PDF (até 5s)
        for i in range(10):
            if os.path.exists(pdf_path):
                print(f"✅ PDF gerado com sucesso: {pdf_path}")
                return pdf_path
            time.sleep(0.5)

        print("❌ PDF não foi detectado após conversão.")
        return None

    except subprocess.TimeoutExpired:
        print("⏰ Tempo limite excedido ao tentar gerar o PDF.")
        return None
    except Exception as e:
        print(f"❌ Erro ao converter Excel para PDF: {e}")
        return None
    
def preparar_planilha_para_pdf(wb, escalas_por_aba=None, escala_padrao=75, print_areas=None):
    """
    Configura as abas para exportação em PDF centralizado.
    """
    if escalas_por_aba is None:
        escalas_por_aba = {}
    if print_areas is None:
        print_areas = {}

    for ws in wb.worksheets:
        escala = escalas_por_aba.get(ws.title, escala_padrao)
        area = print_areas.get(ws.title)

        ws.page_margins = PageMargins(
            left=0.3, right=0.3, top=0.4, bottom=0.4,
            header=0.0, footer=0.0
        )
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.scale = escala
        ws.page_setup.fitToWidth = False
        ws.page_setup.fitToHeight = False

        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True

        if area:
            ws.print_area = area

def adicionar_legenda_layers(ws, legenda_layers, entidades_exemplo, start_row=1, start_col=9):
    import os
    import matplotlib.pyplot as plt
    from matplotlib.patches import Circle, Polygon, Rectangle
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage

    # Estilos
    font_bold   = Font(bold=True, size=12)
    font_normal = Font(size=10)
    align_left  = Alignment(horizontal="left", vertical="center")
    align_center= Alignment(horizontal="center", vertical="center")

    # Cabeçalho
    ws.merge_cells(start_row=start_row, start_column=start_col+1,
                   end_row=start_row,   end_column=start_col+2)
    t = ws.cell(row=start_row, column=start_col+1, value="PROJETO DE SISTEMATIZAÇÃO")
    t.font      = font_bold
    t.alignment = align_center

    os.makedirs("output", exist_ok=True)
    row = start_row + 2
    col_letter = get_column_letter(start_col)

    # calcula largura da célula em pixels (aprox. 7px por caractere)
    cw = ws.column_dimensions[col_letter].width or 8.43
    cell_width_px = int(cw * 7)

    for layer_name in legenda_layers:
        upper = layer_name.upper()
        entidade = entidades_exemplo.get(layer_name)

        # 🔴 LOMBADA: forçado para CIRCLE vermelho
        if "LOMBADA" in upper:
            shape = "CIRCLE"
            cor = (1.0, 0.0, 0.0)

        else:
            cor = legenda_layers[layer_name].get("color")

            # 🔍 Ignora se for branco
            if cor in [None, (1.0, 1.0, 1.0)]:
                continue

            # Tipo de forma
            if upper == "XLEGENDA SISTEMATIZAÇÃO":
                shape = "TRIANGLE"
            else:
                shape = entidade["type"] if entidade else "EMPTY"


        # desenha no Matplotlib num quadro 100×100px
        fig, ax = plt.subplots(figsize=(1,1), dpi=100)
        ax.axis("off"); ax.set_aspect("equal"); ax.set_facecolor("white")

        try:
            if shape == "TRIANGLE":
                patch = Polygon([[0.5,0.1],[0.1,0.9],[0.9,0.9]],
                                closed=True, facecolor=cor, edgecolor="black", linewidth=1.2,
                                transform=ax.transAxes)
                ax.add_patch(patch)

            elif shape == "CIRCLE":
                patch = Circle((0.5,0.5),0.25,
                               facecolor=cor, edgecolor="black", linewidth=1.2,
                               transform=ax.transAxes)
                ax.add_patch(patch)

            elif shape == "EMPTY":
                patch = Rectangle((0.25,0.25),0.5,0.5,
                                  facecolor="white", edgecolor="black", linewidth=1.5,
                                  transform=ax.transAxes)
                ax.add_patch(patch)

            else:
                # LINE/POLYLINE/SOLID → quadrado com cor do layer
                patch = Rectangle((0.25,0.25),0.5,0.5,
                                  facecolor=cor, edgecolor="black", linewidth=1.5,
                                  transform=ax.transAxes)
                ax.add_patch(patch)

        except Exception as e:
            print(f"⚠️ Erro ao desenhar '{layer_name}': {e}")
            plt.close(fig)
            row += 1
            continue

        # salva raw
        raw_path = f"output/mini_{layer_name}_raw.png"
        fig.savefig(raw_path, bbox_inches="tight", pad_inches=0.1, transparent=True)
        plt.close(fig)

        # abre e redimensiona para 30×30
        im = PILImage.open(raw_path).convert("RGBA")
        thumb = im.resize((30,30), PILImage.LANCZOS)

        # cria canvas final com padding transparente à esquerda
        final_path = f"output/mini_{layer_name}.png"
        canvas = PILImage.new("RGBA", (cell_width_px, 30), (0,0,0,0))
        offset_x = max(cell_width_px - 30, 0)
        canvas.paste(thumb, (offset_x, 0), thumb)
        canvas.save(final_path)

        # insere imagem ancorada no canto superior esquerdo da célula
        img = XLImage(final_path)
        img.width  = cell_width_px
        img.height = 30
        ws.add_image(img, f"{col_letter}{row}")

        # escreve o nome ao lado
        c = ws.cell(row=row, column=start_col+1, value=layer_name)
        c.font      = font_normal
        c.alignment = align_left

        row += 1

def redimensionar_imagem(imagem_path, largura, altura):
    try:
        with Image.open(imagem_path) as img:
            resized_img = img.resize((largura, altura), Image.LANCZOS)
            resized_img.save(imagem_path)
            print("✅ Imagem redimensionada para:", resized_img.size)
    except Exception as e:
        print(f"❌ Erro ao redimensionar imagem: {e}")

def limpar_colunas_fora_do_layout(ws, ultima_coluna_valida="K"):
    col_idx = openpyxl.utils.column_index_from_string(ultima_coluna_valida)
    for i in range(col_idx + 1, 100):  # limpa colunas de L até CV
        col = get_column_letter(i)
        if col in ws.column_dimensions:
            del ws.column_dimensions[col]

def limpar_linhas_fora_do_layout(ws, ultima_linha_valida=33):
    for i in range(ultima_linha_valida + 1, 200):  # limpa linhas 34 em diante
        if i in ws.row_dimensions:
            del ws.row_dimensions[i]

from openpyxl.styles import PatternFill

def adicionar_tabela_comprimentos_custom(ws, layer_data, start_row=1, start_col=1):
    # --------------------------
    #   Configurações de estilo
    # --------------------------
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=False)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=False)
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

    # --------------------------
    #   Montar título mesclado
    # --------------------------
    title_cell = ws.cell(row=start_row, column=start_col, value="COMPRIMENTOS POR LAYER")
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 3
    )

    # --------------------------
    #   Cabeçalho
    # --------------------------
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_color = Font(bold=True, color="FFFFFF")

    headers = ["NOME DO LAYER", "QTD", "TOTAL (m)", "MÉDIA (m)"]
    header_row = start_row + 1
    for i, header_text in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i, value=header_text)
        cell.font = header_font_color
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # --------------------------
    #   Inserir dados
    # --------------------------
    data_start_row = header_row + 1
    current_row = data_start_row

    for layer in sorted(layer_data):
        data = layer_data[layer]
        qtd = data.get("qtd", 0)
        total_m = data.get("total", 0.0)
        media_m = total_m / qtd if qtd > 0 else 0.0

        # Coluna 1: NOME DO LAYER
        cell = ws.cell(row=current_row, column=start_col, value=layer)
        cell.font = cell_font
        cell.alignment = left_alignment
        cell.border = thin_border

        # Coluna 2: QTD
        cell = ws.cell(row=current_row, column=start_col + 1, value=qtd)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border

        # Coluna 3: TOTAL (m)
        cell = ws.cell(row=current_row, column=start_col + 2, value=round(total_m, 2))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border

        # Coluna 4: MÉDIA (m)
        cell = ws.cell(row=current_row, column=start_col + 3, value=round(media_m, 2))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border

        current_row += 1
    
def adicionar_tabela_areas_custom(ws, areas_dict, start_row=1, start_col=1):
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_color = Font(bold=True, color="FFFFFF")

    # Título
    ws.cell(row=start_row, column=start_col, value="ÁREAS CALCULADAS").font = title_font
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 1
    )
    ws.cell(row=start_row, column=start_col).alignment = center_alignment

    # Cabeçalho
    headers = ["NOME", "ÁREA (ha)"]
    for i, header_text in enumerate(headers):
        c = ws.cell(row=start_row + 1, column=start_col + i, value=header_text)
        c.font = header_font_color
        c.alignment = center_alignment
        c.fill = header_fill
        c.border = thin_border

    # Inserir dados
    current_row = start_row + 2
    for name, area_m2 in areas_dict.items():
        area_ha = area_m2 / 10000

        c = ws.cell(row=current_row, column=start_col, value=name)
        c.font = cell_font
        c.alignment = left_alignment
        c.border = thin_border

        c = ws.cell(row=current_row, column=start_col + 1, value=round(area_ha, 2))
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        current_row += 1

def parse_talhao_layer_name(layer_name):
    """
    Recebe algo como '06.11.14' e retorna ('06', 11.14).
    Se não houver ponto ou não for possível converter a área, 
    retorna (layer_name, 0.0).
    """
    parts = layer_name.split('.', 1)  # Divide em 2 partes no primeiro ponto
    if len(parts) == 2:
        numero_str, area_str = parts
        numero_str = numero_str.strip()
        try:
            area_ha = float(area_str)
        except ValueError:
            area_ha = 0.0
    else:
        # Se não houver ponto, ou não der para converter, 
        # assume o layer_name inteiro como número e área = 0
        numero_str = layer_name
        area_ha = 0.0

    return numero_str, area_ha

def adicionar_tabela_talhoes_custom(ws, talhoes_dict, start_row=1, start_col=1):
    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

    # Título
    ws.cell(row=start_row, column=start_col, value="TALHÕES").font = title_font
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 2
    )
    ws.cell(row=start_row, column=start_col).alignment = center_alignment

    # Cabeçalho
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_color = Font(bold=True, color="FFFFFF")

    headers = ["Número", "Área (ha)", "Área (alq)*"]
    for i, header_text in enumerate(headers):
        c = ws.cell(row=start_row+1, column=start_col + i, value=header_text)
        c.font = header_font_color
        c.alignment = center_alignment
        c.fill = header_fill
        c.border = thin_border
        
    # Inserir dados
    total_ha = 0.0
    total_alq = 0.0
    row_data_start = start_row + 2

    current_row = row_data_start
    for numero, area_ha in talhoes_dict.items():
        area_alq = area_ha / 2.42

        c = ws.cell(row=current_row, column=start_col, value=numero)
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        c = ws.cell(row=current_row, column=start_col + 1, value=round(area_ha, 2))
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        c = ws.cell(row=current_row, column=start_col + 2, value=round(area_alq, 2))
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        total_ha += area_ha
        total_alq += area_alq
        current_row += 1

    c = ws.cell(row=current_row, column=start_col, value="TOTAL")
    c.font = Font(bold=True, color="FF0000")
    c.alignment = center_alignment
    c.border = thin_border

    c = ws.cell(row=current_row, column=start_col + 1, value=round(total_ha, 2))
    c.font = Font(bold=True)
    c.alignment = center_alignment
    c.border = thin_border

    c = ws.cell(row=current_row, column=start_col + 2, value=round(total_alq, 2))
    c.font = Font(bold=True)
    c.alignment = center_alignment
    c.border = thin_border

    ws.cell(row=current_row + 1, column=start_col + 2, value="*Alqueires Paulistas").alignment = Alignment(horizontal="right")

    ws.column_dimensions[get_column_letter(start_col)].width = 10
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 12
    ws.column_dimensions[get_column_letter(start_col + 2)].width = 12


def gerar_layout_final(dxf_file_path, layer_data, talhoes_dict, legenda_layers, entidades_visiveis, dados):
    # Aqui você usa 'dados' diretamente, sem abrir outro diálogo.
    print("Dados recebidos:", dados)

    def gerar_nome_excel(dxf_file_path, versao_anterior=None):
        nome_dxf = os.path.splitext(os.path.basename(dxf_file_path))[0]
        if versao_anterior is None:
            versao = 0.1
        else:
            try:
                versao = round(float(versao_anterior) + 0.1, 1)
            except ValueError:
                versao = 0.1
        return f"{nome_dxf}_V{versao}.xlsx"

    
    def centralizar_imagem_na_planilha(ws, imagem_path, cell_coord="A5"):
        from openpyxl.utils import get_column_letter

        if not os.path.exists(imagem_path):
            print("❌ Imagem do mapa não foi encontrada.")
            return

        try:
            img = XLImage(imagem_path)
            img.width = 800   # 📏 tamanho VISUAL fixo na célula
            img.height = 575  # 📏 mas a imagem continua HD
            img.anchor = cell_coord
            ws.add_image(img)
            print(f"✅ Imagem em alta resolução inserida na célula {cell_coord}")
        except Exception as e:
            print(f"❌ Erro ao inserir imagem: {e}")

    template_file = resource_path('resources/excel/Planilha_template.xlsx')

    # Usa o diretório definido pelo usuário (via interface), ou fallback para 'output'
    output_dir = dados.get("out_dir", os.path.join(os.path.dirname(__file__), '..', 'output'))
    os.makedirs(output_dir, exist_ok=True)

    output_file = os.path.join(output_dir, gerar_nome_excel(dxf_file_path))

    wb = openpyxl.load_workbook(template_file)
    if "Pagina1" not in wb.sheetnames or "Pagina2" not in wb.sheetnames:
        print("❌ As abas 'Pagina1' ou 'Pagina2' não foram encontradas no template.")
        return

    ws_pagina1 = wb["Pagina1"]
    ws_pagina2 = wb["Pagina2"]

    # Supondo que as funções abaixo já estão definidas em seu projeto:
    # limpar_colunas_fora_do_layout(ws, ultima_coluna_valida)
    # limpar_linhas_fora_do_layout(ws, ultima_linha_valida)
    # preparar_planilha_para_pdf(wb, escalas_por_aba, print_areas)
    # adicionar_legenda_layers(ws, legenda_layers, start_row, start_col)
    # adicionar_tabela_comprimentos_custom(ws, layer_data, start_row, start_col)
    # adicionar_tabela_talhoes_custom(ws, talhoes_dict, start_row, start_col)
    # set_cell_value(ws, cell_coord, value)

    try:
        from ui.layout_generator import (
            limpar_colunas_fora_do_layout, limpar_linhas_fora_do_layout,
            preparar_planilha_para_pdf, adicionar_legenda_layers,
            adicionar_tabela_comprimentos_custom, adicionar_tabela_talhoes_custom, set_cell_value
        )
    except ImportError as e:
        print("Erro ao importar funções auxiliares:", e)
        return

    limpar_colunas_fora_do_layout(ws_pagina1, "K")
    limpar_linhas_fora_do_layout(ws_pagina1, 40)
    limpar_colunas_fora_do_layout(ws_pagina2, "J")
    limpar_linhas_fora_do_layout(ws_pagina2, 40)

    preparar_planilha_para_pdf(
        wb,
        escalas_por_aba={"Pagina1": 75, "Pagina2": 85},
        print_areas={"Pagina1": "A1:K40", "Pagina2": "A1:J40"}
    )

    ws_pagina1.merge_cells("H36:I36")
    ws_pagina2.merge_cells("F35:J35")

    # Logo da Cevasa
    img_cevasa_path = resource_path("resources/images/logo.png")
    redimensionar_imagem(img_cevasa_path, 95, 40)
    img_cevasa = XLImage(img_cevasa_path)
    img_cevasa.anchor = "A34"
    ws_pagina2.add_image(img_cevasa)
    ws_pagina1.column_dimensions["K"].width = 36

    # Rosa dos ventos
    img_rosa_path_1 = resource_path("resources/images/rosa_dos_ventos.png")
    redimensionar_imagem(img_rosa_path_1, 110, 110)
    img_rosa_path_2 = resource_path("resources/images/rosa_dos_ventos.png")
    redimensionar_imagem(img_rosa_path_2, 100, 90)

    ws_pagina1.merge_cells("K31:K34")
    ws_pagina2.merge_cells("I30:J33")

    img_final_rosa_1 = os.path.join("output", "rosa_dos_ventos_pagina1.png")
    img_final_rosa_2 = os.path.join("output", "rosa_dos_ventos_pagina2.png")
    gerar_imagem_centrada(img_rosa_path_1, 252, 110, img_final_rosa_1)
    inserir_imagem(ws_pagina1, img_final_rosa_1, "K31")
    gerar_imagem_centrada(img_rosa_path_2, 170, 90, img_final_rosa_2)
    inserir_imagem(ws_pagina2, img_final_rosa_2, "I30")
    ws_pagina1.column_dimensions["K"].width = 36

    img_cevasa = XLImage(resource_path("resources/images/logo.png"))
    img_cevasa.anchor = "A35"
    ws_pagina1.add_image(img_cevasa)

    # Preencher informações na planilha
    set_cell_value(ws_pagina1, "I31", dados['parc'])         
    set_cell_value(ws_pagina1, "J32", dados['data_atual'])   
    set_cell_value(ws_pagina1, "I33", dados['distancia'])      
    set_cell_value(ws_pagina1, "I34", dados['area_cana'])      
    set_cell_value(ws_pagina1, "J34", dados['nova_versao'])    
    set_cell_value(ws_pagina1, "I32", dados['escala'])         
    set_cell_value(ws_pagina1, "B36", dados['propriedade'])    
    set_cell_value(ws_pagina1, "E36", dados['mun_est'])       
    set_cell_value(ws_pagina1, "H36", dados['desenhista'])    

    set_cell_value(ws_pagina2, "G30", dados['parc'])        
    set_cell_value(ws_pagina2, "H31", dados['data_atual'])     
    set_cell_value(ws_pagina2, "G32", dados['distancia'])      
    set_cell_value(ws_pagina2, "G33", dados['area_cana'])     
    set_cell_value(ws_pagina2, "H33", dados['nova_versao']) 
    set_cell_value(ws_pagina2, "G31", dados['escala'])        
    set_cell_value(ws_pagina2, "B35", dados['propriedade'])    
    set_cell_value(ws_pagina2, "C35", dados['mun_est'])        
    set_cell_value(ws_pagina2, "F35", dados['desenhista'])   

    
    # 📦 Agrupa entidades visíveis por layer
    # 📦 Mapeia entidades visíveis por layer
    entidades_por_layer = {}
    for ent in entidades_visiveis:
        layer = ent["layer"]
        entidades_por_layer.setdefault(layer, []).append(ent)

    def is_branco(cor):
        if not isinstance(cor, (tuple, list)) or len(cor) != 3:
            return True
        return all(c >= 0.95 for c in cor)

    # 🔽 Refiltra legenda: só layers com pelo menos 1 entidade NÃO-texto e cor visível
    legenda_filtrada = dict(sorted([
        (layer, info)
        for layer, info in legenda_layers.items()
        if any(e["type"] not in ["TEXT", "MTEXT"] for e in entidades_por_layer.get(layer, []))
        and not is_branco(info.get("color"))
    ], key=lambda x: x[0]))

    exemplos_legenda = {}
    for ent in entidades_visiveis:
        layer = ent["layer"]
        if layer in legenda_filtrada and layer not in exemplos_legenda:
            exemplos_legenda[layer] = ent
    
    print("📊 Layers com comprimento calculado:", layer_data.keys())
    print("✅ Layers marcados pelo usuário:", dados["entidades_exemplo"].keys())

    print(">>> TALHÕES DICT:", talhoes_dict)

    adicionar_tabela_talhoes_custom(ws_pagina2, talhoes_dict, start_row=4, start_col=7)
    adicionar_legenda_layers(ws_pagina1, legenda_filtrada, exemplos_legenda, start_row=4, start_col=9)

    def area_da_entidade(ent):
        if ent["type"] in ["POLYLINE", "LWPOLYLINE", "SOLID"]:
            pts = ent.get("points", [])
            if len(pts) >= 3:
                return abs(Polygon(pts).area)
        if ent["type"] == "CIRCLE":
            return math.pi * (ent["radius"] ** 2)
        return 0.0

    def area_por_layer(dxf_entities):
        areas = defaultdict(float)
        for ent in dxf_entities:
            a = area_da_entidade(ent)
            if a > 0:
                areas[ent["layer"]] += a
        return areas

    def diagnostico_areas_por_layer(dxf_entities):
        areas = defaultdict(float)
        for ent in dxf_entities:
            a = area_da_entidade(ent)
            if a > 0:
                areas[ent["layer"]] += a
        
        print("\n=== DIAGNÓSTICO DAS ÁREAS POR LAYER ===")
        for layer, total_area in sorted(areas.items()):
            print(f"Layer '{layer}': {total_area:.2f} m² → {total_area / 10000:.2f} ha")
        print("========================================\n")
        
        area_preto = areas.get("0", 0.0)
        area_talhoes = areas.get("TALHÕES", 0.0)
        area_carreador = area_preto - area_talhoes
        
        print(f"Área TOTAL layer '0': {area_preto:.2f} m² → {area_preto/10000:.2f} ha")
        print(f"Área TOTAL layer 'TALHÕES': {area_talhoes:.2f} m² → {area_talhoes/10000:.2f} ha")
        print(f"Área do CARREADOR calculada: {area_carreador:.2f} m² → {area_carreador/10000:.2f} ha")
        
        return areas, area_carreador

    areas, area_carreador = diagnostico_areas_por_layer(entidades_visiveis)

    areas_para_exibir = {
        "Layer '0' (preto)": areas.get("0", 0.0),
        "Layer 'TALHÕES'": areas.get("TALHÕES", 0.0),
        "Área Carreador (calculada)": area_carreador
    }

    adicionar_tabela_areas_custom(ws_pagina2, areas_para_exibir, start_row=21, start_col=2)
    adicionar_tabela_comprimentos_custom(ws_pagina2, layer_data, start_row=4, start_col=2)

    image_path = os.path.join("input", "mapa.png")

    if os.path.exists(image_path):
        try:
            centralizar_imagem_na_planilha(ws_pagina1, image_path, cell_coord="A05")
            print("✅ Imagem 'mapa.png' adicionada na aba 'Pagina1'.")
        except Exception as e:
            print(f"❌ Erro ao inserir imagem 'mapa.png': {e}")
    else:
        print("❌ Imagem do mapa não foi encontrada no caminho:", image_path)

    wb.save(output_file)
    print(f"✅ Planilha salva como '{output_file}'.")

    pdf_path = converter_excel_para_pdf_com_libreoffice(output_file)
    
    return pdf_path if os.path.exists(pdf_path) else None
